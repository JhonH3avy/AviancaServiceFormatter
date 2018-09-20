from os import path
from openpyxl import load_workbook
from Service import Service
import services_to_excel
import sys

OK = 0
FILE_LOADING_ERROR = 1
FILE_FORMAT_ERROR = 2
FILE_ACCESS_PERMISSION_ERROR = 3

def main():
    """Loads and modifies a sample spreadsheet in Excel."""
    flags = {}
    flag = None
    for arg in sys.argv:
        if '-' in arg:
            flag = arg
        else:
            if flag:
                flags[flag] = arg
                flag = None

    
    if '-p' in flags:
        workbook_path = flags['-p']
        print('Using custom path')
    else:
        workbook_path = path.join(path.dirname(__file__), 'service_data.xlsx')

    if '-o' in flags:
        output_path = flags['-o']
    else:
        output_path = path.dirname(__file__) + 'services.xlsx'

    try:
        workbook = load_workbook(workbook_path, read_only=True)
        print('The file has been loaded successfully')
    except FileNotFoundError:
        print('The file has not been found in ' + workbook_path);
        sys.exit(FILE_LOADING_ERROR)
    except FileExistsError:
        print('The file ' + workbook_path + ' doesn\'t exist')
        sys.exit(FILE_LOADING_ERROR)


    sheet_names = workbook.sheetnames
    servicesToParse = None
    for sheet in sheet_names:
        print('Getting services of sheet ' + sheet)
        servicesToParse = process_sheet(sheet, workbook, servicesToParse)

    print(str(len(servicesToParse)) + ' services has been created from file ' + workbook_path)
    normalized_excel = services_to_excel.convert_to_excel(servicesToParse)

    try:
        normalized_excel.save(output_path)
        print('Excel file created successfully')
    except PermissionError:
        print('Program do not have permission to access ' + output_path)
        sys.exit(FILE_ACCESS_PERMISSION_ERROR)
    print('Job Done...')
    sys.exit(OK)    



def process_sheet(sheet, workbook, services = []):   
    if services is None:
        services = []
    if not is_valid_sheet_name(sheet):
        print(sheet + ' is not a valid sheet name.')
        return services

    sheet_ranges = workbook[sheet]
    
    columnDict = get_columns_index(sheet_ranges)

    try:
        assert sheet_ranges['A5'].value is not None
        local = sheet_ranges['A5'].value
        localData = local.split(':')
    except AssertionError:
        print('File is not in correct format. Cell "A5" has no destination nor origin information')
        sys.exit(FILE_FORMAT_ERROR)

    origin = ''
    destination = ''

    if 'ORG' in localData[0]:
        origin = localData[1].strip()
    else:
        destination = localData[1].strip()


    max_col = sheet_ranges.rows.gi_frame.f_locals['max_col']
    max_row = sheet_ranges.rows.gi_frame.f_locals['max_row']

    for row in sheet_ranges.iter_rows(min_row=9, max_col=max_col, max_row=max_row):
        fields = {}
        for cell in row: 
            if hasattr(cell,'column') and  cell.column in columnDict.keys():
                fields[columnDict[cell.column]] = cell.value
            else:
                continue
        
        if len(fields) == 0:
            continue

        if origin:
            fields['origin'] = origin
        else:
            fields['destination'] = destination

        service = Service(fields)
        if service.is_valid_service():
            result = [serviceToFilter for serviceToFilter in services if service.paxName == serviceToFilter.paxName]
            if result:
                indx = services.index(result[0])
                if result[0].is_arriving():
                    service_with_connection = put_in_connection_of_service(services[indx], service)
                else:
                    service_with_connection = put_in_connection_of_service(service, services[indx])
                services.remove(result[0])
                services.append(service_with_connection)
            else:
                services.append(service)
        
    print('Services for ' + sheet + ' has been processed')   
    return services
    


def is_valid_sheet_name(sheetName):
    keys = ['SALIDAS', 'SALIENDO', 'LLEGADAS', 'LLEGANDO']
    for key in keys:
        if key in sheetName.upper():
            return True
    return False
            

def get_columns_index(sheet_ranges):
    columnIdx = {}
    max_col = sheet_ranges.rows.gi_frame.f_locals['max_col']
    for i in range(7,9):
        for j in range(1,(max_col+1)):
            cell = sheet_ranges.cell(row=i, column=j)
            columnName = get_dict_equivalent(cell.value)
            if columnName:
                columnIdx[j] = columnName
    return columnIdx

def get_dict_equivalent(columnName):
    dict = {
        'SSR_SCARR':'carrier',
        'SSR_SORG':'origin',
        'SSR_SDST':'destination',
        'SSR_FTNR':'flightNumber',
        'SSR_DEPDATE':'departure',
        'SSR_ARRDATE':'arrival',
        'SSR_PAXNAME':'paxName',
        'WCOB':'WCOB',
        'WCMP':'WCMP',
        'WCHS':'WCHS',
        'WCHR':'WCHR',
        'WCHC':'WCHC',
        'WCBW':'WCBW',
        'WEAP':'WEAP'
    }
    if columnName in dict.keys():
        return dict[columnName]
    else:
        return None   


def put_in_connection_of_service(service, connection):
    service.flightConnectionNumber = connection.flightNumber
    service.endZone = 'SALAS'
    return service


if __name__ == '__main__':
    main()
