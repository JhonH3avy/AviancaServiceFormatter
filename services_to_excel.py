from openpyxl import Workbook

def convert_to_excel(services = []):
    wb = Workbook()
    ws = wb.active
    ws.title = "Servicios Avianca"
    ws['A1'] = 'flightId'
    ws['B1'] = 'airline'
    ws['C1'] = 'originCity'
    ws['D1'] = 'destinyCity'
    ws['E1'] = 'startDate'
    ws['F1'] = 'flightConnectionId'
    ws['G1'] = 'paxName'
    ws['H1'] = 'paxReservationNumber'
    ws['I1'] = 'startZone'
    ws['J1'] = 'endZone'
    ws['K1'] = 'connectionGate'
    ws['L1'] = 'serviceType'
    ws['M1'] = 'reserved'
    ws['N1'] = 'authorizedBy'

    total_services = len(services)
    for i in range(0,total_services):
        ws['A' + str(i+2)] = services[i].flightNumber
        ws['B' + str(i+2)] = services[i].airline
        ws['C' + str(i+2)] = services[i].origin
        ws['D' + str(i+2)] = services[i].destination
        ws['E' + str(i+2)] = services[i].startDate
        ws['F' + str(i+2)] = services[i].flightConnectionNumber
        ws['G' + str(i+2)] = services[i].paxName
        ws['H' + str(i+2)] = services[i].paxReservationNumber
        ws['I' + str(i+2)] = services[i].startZone
        ws['J' + str(i+2)] = services[i].endZone
        ws['K' + str(i+2)] = services[i].connectionGate
        ws['L' + str(i+2)] = services[i].serviceType
        ws['M' + str(i+2)] = services[i].reserved
        ws['N' + str(i+2)] = services[i].authorizedBy


    return wb